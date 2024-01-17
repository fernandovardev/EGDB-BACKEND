import pandas as pd
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Border, Side, Alignment
from openpyxl.styles import numbers
from core.gasdata.DB.db_config import super_stations
import io
##########################################################################################################################
############################################ GENERACIÓN DE REPORTE DIARIO #################################################
##########################################################################################################################
def generate_daily_report_df(sorted_df, efectivo_df, date_filter):
    station_data = sorted_df[sorted_df['FECHA'] == date_filter].groupby('ESTACION')
    efectivo_data = efectivo_df[efectivo_df['FECHA'] == date_filter] .drop(columns=['FECHA'])

    # Process station data
    processed_station_data = {}
    for station_id, group in station_data:
        group = group.drop(columns=['FECHA', 'ESTACION', 'VOLUMEN'])
        group = group[['PRODUCTO', 'LTSVENTA', 'FISICA', 'LTSMOV']]
        group['PRODUCTO'] = group['PRODUCTO'].replace({"0001": "DIESEL", 
                                                       "0002": "MAGNA", 
                                                       "0003": "PREMIUM"})
        group.rename(columns={
            'LTSVENTA': 'VENTA',
            'LTSMOV': 'DESC.',
            'FISICA': 'MEDIDA',
        }, inplace=True)
        processed_station_data[station_id] = group
    return processed_station_data, efectivo_data

def output_to_excel_daily_report(sorted_df, efectivo_df,date_filter):
    station_data = sorted_df[sorted_df['FECHA'] == date_filter].groupby('ESTACION')
    efectivo_data = efectivo_df[efectivo_df['FECHA'] == date_filter]
    # Create a new Excel workbook and select the active sheet
    wb = Workbook()
    ws = wb.active
    ws.title = "Station Data"
    
    # Add three empty rows at the beginning
    ws.insert_rows(1, amount=3)
    
    # Add the report header with the date
    report_header = f"REPORTE DE VENTAS POR ESTACION CON FECHA: {date_filter.strftime('%d/%m/%Y')}"
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=4)
    header_cell = ws.cell(row=2, column=1, value=report_header)
    header_cell.font = Font(bold=True)
    header_cell.alignment = Alignment(horizontal='center', vertical='center')

    # Define the styles
    bold_font = Font(bold=True, color='FF0000FF')
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    center_aligned_text = Alignment(horizontal='center', vertical='center')
    number_format = numbers.BUILTIN_FORMATS[4]  # "0.00"
    money_format=numbers.BUILTIN_FORMATS[7] #'"$"#,##0_);("$"#,##0)',
    # Set standard column width and calculate maximum width required for headers
    column_widths = {'PRODUCTO': 30, 'VENTA': 10, 'MEDIDA': 10, 'DESC.': 10}
    # Set the headers
    headers = ['PRODUCTO', 'VENTA', 'MEDIDA', 'DESC.']

    # Initialize starting position

    start_row = 4  # Start from the second row to leave space for station names
    start_column = 1
    max_row = 1  # Tracks the maximum number of rows used in the last printed table
    table_spacing = 1  # Space between tables
    num_columns_in_grid = 3
    max_col_widths = 25

    for idx, (station_id, group) in enumerate(station_data):
        group = group.drop(columns=['FECHA','ESTACION','VOLUMEN'])
        group = group[['PRODUCTO','LTSVENTA','FISICA','LTSMOV']]
        group.loc[group.PRODUCTO == "0001", 'PRODUCTO'] = 'DIESEL'
        group.loc[group.PRODUCTO == "0002", 'PRODUCTO'] = 'MAGNA'
        group.loc[group.PRODUCTO == "0003", 'PRODUCTO'] = 'PREMIUM'
        group.rename(columns={

        'LTSVENTA': 'VENTA',
        'LTSMOV': 'DESC.',
        'FISICA': 'MEDIDA',

    }, inplace=True)
        # Calculate starting positions
        current_col = ((idx % num_columns_in_grid) * (len(headers) + table_spacing)) + 1
        # Set the starting row for the first set of tables to be row 2 to accommodate station names
        if idx < num_columns_in_grid:
            current_row = start_row
        else:
            current_row = ((idx // num_columns_in_grid) * (max_row + 2 + table_spacing)) + start_row

        # Write the station name above the headers
        ws.merge_cells(start_row=current_row - 1, start_column=current_col, 
                    end_row=current_row - 1, end_column=current_col + len(headers) - 1)
        station_cell = ws.cell(row=current_row - 1, column=current_col, value=f'ÁGUILA {station_id}')
        station_cell.font = bold_font
        station_cell.alignment = center_aligned_text

        # Headers
        for i, header in enumerate(headers, start=0):
            cell = ws.cell(row=current_row, column=current_col + i)
            cell.value = header
            cell.font = bold_font
            cell.alignment = center_aligned_text
            cell
            ws.column_dimensions[get_column_letter(current_col + i)].width = column_widths[header]

        # Data
        for r, data_row in enumerate(group.itertuples(index=False), start=current_row + 1):
            for c, value in enumerate(data_row, start=0) :
                cell = ws.cell(row=r, column=current_col + c)
                cell.value = value
                cell.border = thin_border
                cell.alignment = center_aligned_text
                ws.column_dimensions[get_column_letter(current_col + c)].width = max_col_widths
                if c > 0:  # Apply number formatting to non-'PRODUCTO' columns
                    cell.number_format = number_format
        
        # After the data, add the extra row with "PANAMERICANO" and "SUPER" if applicable
        efectivo_row = efectivo_data[efectivo_data['ESTACION'] == station_id]
        print(efectivo_row)
        if not efectivo_row.empty:
            efectivo_row = efectivo_row.iloc[0]
            
            # Insert an empty row for spacing if necessary
            r += 1

            # Add PANAMERICANO row
            panam_cell = ws.cell(row=r + 1, column=current_col)
            panam_cell.value = 'PANAMERICANO'
            panam_value_cell = ws.cell(row=r + 1, column=current_col + 1)
            panam_value_cell.value = efectivo_row['PANAMERICANO']
            panam_cell.border = panam_value_cell.border = thin_border
            panam_cell.alignment = panam_value_cell.alignment = center_aligned_text
            panam_value_cell.number_format = money_format
            r += 1  # Increment to account for the PANAMERICANO row
            
            # Add SUPER row if applicable
            if station_id in super_stations and 'SUPER' in efectivo_row:
                super_cell = ws.cell(row=r + 1, column=current_col)
                super_cell.value = 'SUPER'
                super_value_cell = ws.cell(row=r + 1, column=current_col + 1)
                super_value_cell.value = efectivo_row['SUPER']
                super_cell.border = super_value_cell.border = thin_border
                super_cell.alignment = super_value_cell.alignment = center_aligned_text
                super_value_cell.number_format = money_format
                r += 1  # Increment to account for the SUPER row
        else:
            # Insert an empty row for spacing if necessary
            r += 1

            # Add PANAMERICANO row
            panam_cell = ws.cell(row=r + 1, column=current_col)
            panam_cell.value = 'PANAMERICANO'
            panam_value_cell = ws.cell(row=r + 1, column=current_col + 1)
            panam_value_cell.value = 0
            panam_cell.border = panam_value_cell.border = thin_border
            panam_cell.alignment = panam_value_cell.alignment = center_aligned_text
            panam_value_cell.number_format = money_format
            r += 1  # Increment to account for the PANAMERICANO row
        # Update max_row with the number of rows in the current table if it's greater than the last
        max_row = max(max_row, r - current_row)

        # Update start_column for the next station; reset if at the end of a grid row
        if (idx + 1) % num_columns_in_grid == 0:
            start_column = 1  # Reset to first column
        else:
            start_column += len(headers) + table_spacing

        # Set the calculated column widths after populating the sheet
        for i, header in enumerate(headers, start=1):
            ws.column_dimensions[get_column_letter(i)].width = max_col_widths

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
    return output
